class ThongKeService:
    COT_HEADERS = ["Bien so", "Loai xe", "Vi tri", "Gio vao", "Gio ra", "So phut gui", "Thanh tien"]
    COT_KEYS = ["bien_so", "loai_xe", "vi_tri", "gio_vao", "gio_ra", "so_phut_gui", "thanh_tien"]

    def xuat_excel(self, file_path, du_lieu_doanh_thu, du_lieu_gui_lau):
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        self._ghi_sheet(wb.active, "Doanh thu", du_lieu_doanh_thu, True, openpyxl, Font, Alignment, PatternFill, Border, Side)
        ws2 = wb.create_sheet()
        self._ghi_sheet(ws2, "Xe gui lau nhat", du_lieu_gui_lau, False, openpyxl, Font, Alignment, PatternFill, Border, Side)
        wb.save(file_path)

    def _ghi_sheet(self, ws, tieu_de, du_lieu, hien_tong_doanh_thu, openpyxl, Font, Alignment, PatternFill, Border, Side):
        ws.title = tieu_de
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="176BFF", end_color="176BFF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.COT_HEADERS))
        ws.cell(row=1, column=1, value=tieu_de).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        for col_idx, header in enumerate(self.COT_HEADERS, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, ban_ghi in enumerate(du_lieu, 4):
            for col_idx, key in enumerate(self.COT_KEYS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=ban_ghi.get(key, ""))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        if hien_tong_doanh_thu:
            self._ghi_tong_doanh_thu(ws, du_lieu, len(du_lieu) + 5, Font, Alignment, PatternFill, thin_border)

        for col_idx in range(1, len(self.COT_HEADERS) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    def _ghi_tong_doanh_thu(self, ws, du_lieu, dong_tong, Font, Alignment, PatternFill, thin_border):
        tong_doanh_thu = sum(self._lay_gia_tri_tien(ban_ghi.get("thanh_tien", 0)) for ban_ghi in du_lieu)
        tong_doanh_thu_text = f"{tong_doanh_thu:,}".replace(",", ".") + " d"
        ws.merge_cells(start_row=dong_tong, start_column=1, end_row=dong_tong, end_column=len(self.COT_HEADERS) - 1)
        label_cell = ws.cell(row=dong_tong, column=1, value="TONG DOANH THU")
        value_cell = ws.cell(row=dong_tong, column=len(self.COT_HEADERS), value=tong_doanh_thu_text)

        for cell in (label_cell, value_cell):
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        for col_idx in range(2, len(self.COT_HEADERS)):
            cell = ws.cell(row=dong_tong, column=col_idx)
            cell.fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
            cell.border = thin_border

    def _lay_gia_tri_tien(self, gia_tri):
        if isinstance(gia_tri, int):
            return gia_tri
        chuoi_so = "".join(ky_tu for ky_tu in str(gia_tri) if ky_tu.isdigit())
        return int(chuoi_so) if chuoi_so else 0
