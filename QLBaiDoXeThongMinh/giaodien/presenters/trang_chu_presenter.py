"""
Tầng Presenter
Cầu nối: Lấy dữ liệu từ Service -> đẩy vào View: Nhận signal từ view -> xử lý
nguyên tắc là View và Service không biết gì nhau


"""
from views.trang_chu import ITrangChuView
from views.check_in import CheckInDialog
from presenters.checkin_presenter import CheckInPresenter
from PyQt5 import uic
from PyQt5.QtWidgets import (
    QDialog, QMainWindow, QMessageBox, QPushButton,
    QTableWidgetItem, QListWidget, QListWidgetItem,
    QFileDialog, QVBoxLayout,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QTextDocument
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
import os
import sys




DESIGNER_DIR = os.path.join(os.path.dirname(__file__), "..", "designer")
PROJECT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if PROJECT_DIR not in sys.path:
    sys.path.append(PROJECT_DIR)


from xulyhethong.quan_ly_bai_xe import QuanLyBaiXe
from doituong.loi_ngoai_le import VehicleNotFoundError




class TrangChuPresenter:
    def __init__(self, view:ITrangChuView):
        self._view = view
        self._windows = []
        self._bai_xe = QuanLyBaiXe()
        self._noi_su_kien()


    def _noi_su_kien(self):
        self._view.yeu_cau_checkin.connect(self.mo_man_hinh_check_in)
        self._view.yeu_cau_checkout.connect(self.mo_man_hinh_check_out)
        self._view.yeu_cau_quan_ly_bai_xe.connect(self.mo_man_hinh_quan_ly_bai_xe)
        self._view.yeu_cau_thong_ke_chi_tiet.connect(self.mo_man_hinh_thong_ke_chi_tiet)
        self._view.yeu_cau_so_do.connect(self.mo_man_hinh_so_do)


    def load(self):
        self._cap_nhat_dashboard()


    def mo_man_hinh_check_in(self):
        dialog = CheckInDialog()
        presenter = CheckInPresenter(dialog, self._bai_xe, self._cap_nhat_dashboard)
        dialog.exec_()


    # ----------------------------------------------------------------
    # Task 1: Checkout – hiện danh sách gợi ý xe đang gửi
    # ----------------------------------------------------------------
    def mo_man_hinh_check_out(self):
        dialog = self._load_dialog("check_out.ui")

        # Thêm QListWidget gợi ý ngay bên dưới ô nhập biển số
        if hasattr(dialog, "txt_bien_so"):
            list_goi_y = QListWidget(dialog)
            list_goi_y.setStyleSheet(
                "QListWidget { background: #FFFFFF; border: 1px solid #C8D2DE;"
                " border-radius: 6px; font-size: 14px; color: #050505; }"
                "QListWidget::item { padding: 8px 12px; }"
                "QListWidget::item:hover { background: #E6EDF6; }"
                "QListWidget::item:selected { background: #176BFF; color: #FFFFFF; }"
            )
            list_goi_y.setMaximumHeight(200)

            # Chèn list widget vào layout sau txt_bien_so
            card_layout = dialog.txt_bien_so.parent().layout()
            if card_layout is None:
                # Fallback: tìm card layout
                card_layout = dialog.findChild(QVBoxLayout, "cardLayout")
            if card_layout is not None:
                # Tìm vị trí formLayout trong cardLayout và chèn list sau nó
                idx = -1
                for i in range(card_layout.count()):
                    item = card_layout.itemAt(i)
                    if item and item.layout() and item.layout().objectName() == "formLayout":
                        idx = i
                        break
                if idx >= 0:
                    card_layout.insertWidget(idx + 1, list_goi_y)
                else:
                    card_layout.insertWidget(1, list_goi_y)

            danh_sach_xe = self._bai_xe.lay_danh_sach_xe_dang_gui()

            def cap_nhat_goi_y(text):
                list_goi_y.clear()
                tu_khoa = text.strip().upper()
                for xe in danh_sach_xe:
                    bien_so = xe["bien_so"]
                    loai_xe = xe.get("loai_xe", "")
                    vi_tri = xe.get("vi_tri", "")
                    if not tu_khoa or tu_khoa in bien_so.upper():
                        item = QListWidgetItem(f"{bien_so}  |  {loai_xe}  |  Vị trí: {vi_tri}")
                        item.setData(Qt.UserRole, bien_so)
                        list_goi_y.addItem(item)
                list_goi_y.setVisible(list_goi_y.count() > 0)

            def chon_goi_y(item):
                bien_so = item.data(Qt.UserRole)
                dialog.txt_bien_so.setText(bien_so)
                list_goi_y.hide()

            dialog.txt_bien_so.textChanged.connect(cap_nhat_goi_y)
            list_goi_y.itemClicked.connect(chon_goi_y)

            # Hiện toàn bộ danh sách ban đầu
            cap_nhat_goi_y("")

        if hasattr(dialog, "btn_checkout"):
            dialog.btn_checkout.clicked.connect(lambda: self._xu_ly_check_out(dialog))
        dialog.exec_()


    # ----------------------------------------------------------------
    # Task 9: 2 nút xe đang gửi và lịch sử trỏ đúng tab
    # ----------------------------------------------------------------
    def mo_man_hinh_quan_ly_bai_xe(self):
        self._mo_man_hinh_quan_ly(tab_index=0)


    def _mo_man_hinh_quan_ly(self, tab_index=0):
        window = QMainWindow()
        uic.loadUi(os.path.join(DESIGNER_DIR, "quan_ly_bai_xe.ui"), window)
        self._cap_nhat_bang_quan_ly(window)

        # Đảm bảo tab được set đúng
        if hasattr(window, "tabWidget"):
            window.tabWidget.setCurrentIndex(tab_index)

        # Task 3: Connect nút Xuất Excel
        if hasattr(window, "btn_xuat_excel"):
            window.btn_xuat_excel.clicked.connect(lambda: self._xuat_excel_thong_ke(window))

        window.show()
        self._windows.append(window)


    # ----------------------------------------------------------------
    # Task 2: Sơ đồ bãi xe – nút quay về trang chủ
    # ----------------------------------------------------------------
    def mo_man_hinh_so_do(self, tang=0):
        window = QMainWindow()
        uic.loadUi(os.path.join(DESIGNER_DIR, "so_do.ui"), window)
        self._khoi_tao_so_do(window)
        window.tabWidget.setCurrentIndex(tang)

        # Connect nút quay về
        if hasattr(window, "btn_quay_ve"):
            window.btn_quay_ve.clicked.connect(window.close)

        window.show()
        self._windows.append(window)


    def mo_man_hinh_thong_ke_chi_tiet(self):
        self._mo_man_hinh_quan_ly(tab_index=1)


    def mo_man_hinh_thong_tin_xe_ra(self):
        dialog = self._load_dialog("thong_tin_xe_ra.ui")
        dialog.exec_()


    def _load_dialog(self, ui_file):
        dialog = QDialog()
        uic.loadUi(os.path.join(DESIGNER_DIR, ui_file), dialog)
        return dialog


    def _khoi_tao_so_do(self, window):
        self._ve_tang(window, window.gridLayout, self._bai_xe.lay_du_lieu_theo_loai_slot("XeMay"))
        self._ve_tang(window, window.gridLayout_2, self._bai_xe.lay_du_lieu_theo_loai_slot("OTo"))


    def _ve_tang(self, window, layout, du_lieu_tang):
        for row_group, hang_slot in enumerate(du_lieu_tang):
            for index, slot in enumerate(hang_slot):
                button = self._tao_nut_vi_tri(window, slot)
                hang = row_group
                cot = index + index // 2
                layout.addWidget(button, hang, cot)


    def _tao_nut_vi_tri(self, window, slot):
        ma_vi_tri = slot["ma_vi_tri"]
        con_trong = slot["trang_thai"] == "Trống"
        button = QPushButton("\n".join(ma_vi_tri))
        button.setProperty("class", "SlotBtn")
        button.setMinimumSize(58, 72)
        button.setStyleSheet(
            "background-color: #09C56B; color: #050505; border-radius: 8px; font-weight: 900;"
            if con_trong
            else "background-color: #FF3238; color: #050505; border-radius: 8px; font-weight: 900;"
        )
        trang_thai = "Còn trống" if con_trong else f"Đã có xe\nBiển số: {slot['bien_so']}"
        button.clicked.connect(
            lambda checked=False, vt=ma_vi_tri, tt=trang_thai: QMessageBox.information(
                window,
                "Thông tin vị trí",
                f"Vị trí {vt}: {tt}",
            )
        )
        return button


    def _cap_nhat_dashboard(self):
        self._view.hien_so_xe(
            self._bai_xe.lay_so_xe_theo_loai("XeMay"),
            self._bai_xe.lay_so_xe_theo_loai("OTo"),
        )
        self._view.hien_doanh_thu(self._bai_xe.lay_doanh_thu_theo_ngay())
        self._view.hien_xe_cho(
            self._bai_xe.lay_so_xe_cho_theo_loai("XeMay"),
            self._bai_xe.lay_so_xe_cho_theo_loai("OTo"),
        )
        if hasattr(self._view, "hien_ty_le_lap_day"):
            self._view.hien_ty_le_lap_day(
                self._bai_xe.lay_tong_so_xe_dang_gui(),
                self._bai_xe.lay_tong_so_o(),
                self._bai_xe.lay_ty_le_lap_day(),
            )
        self._cap_nhat_cac_bang_quan_ly()


    def _xu_ly_check_out(self, dialog):
        bien_so = dialog.txt_bien_so.text().strip() if hasattr(dialog, "txt_bien_so") else ""
        if not bien_so:
            QMessageBox.warning(dialog, "Lỗi", "Vui lòng nhập biển số xe")
            return


        try:
            bien_lai = self._bai_xe.check_out_va_lap_bien_lai(bien_so)
        except VehicleNotFoundError as error:
            QMessageBox.warning(dialog, "Lỗi", str(error))
            return


        self._cap_nhat_dashboard()
        dialog.accept()
        self._mo_thong_tin_xe_ra(bien_lai)


    # ----------------------------------------------------------------
    # Task 11: In vé xe ra
    # ----------------------------------------------------------------
    def _mo_thong_tin_xe_ra(self, bien_lai):
        xe_ra = bien_lai["xe_ra"]
        ma_vi_tri = bien_lai["ma_vi_tri"]
        thoi_gian_ra = bien_lai["thoi_gian_ra"]
        tien_gui = bien_lai["tien_gui"]
        so_phut = int(bien_lai["so_phut_gui"])

        dialog = self._load_dialog("thong_tin_xe_ra.ui")
        if hasattr(dialog, "lbl_bien_so"):
            dialog.lbl_bien_so.setText(xe_ra.bien_so)
        if hasattr(dialog, "lbl_loai_xe"):
            dialog.lbl_loai_xe.setText("Xe máy" if xe_ra.loai_xe == "XeMay" else "Ô tô")
        if hasattr(dialog, "lbl_vi_tri_do"):
            dialog.lbl_vi_tri_do.setText(ma_vi_tri)
        if hasattr(dialog, "lbl_gio_vao"):
            dialog.lbl_gio_vao.setText(xe_ra.thoi_gian_vao.strftime("%d/%m/%Y %H:%M"))
        if hasattr(dialog, "lbl_gio_ra"):
            dialog.lbl_gio_ra.setText(thoi_gian_ra.strftime("%d/%m/%Y %H:%M"))
        if hasattr(dialog, "lbl_thoi_gian_gui"):
            dialog.lbl_thoi_gian_gui.setText(f"{so_phut // 60} giờ {so_phut % 60} phút")
        if hasattr(dialog, "lbl_thanh_tien"):
            dialog.lbl_thanh_tien.setText(f"{tien_gui:,}".replace(",", ".") + " đ")

        # Lưu biên lai để dùng cho in vé
        dialog._bien_lai = bien_lai

        # Connect nút IN VÉ
        if hasattr(dialog, "btn_in_ve"):
            dialog.btn_in_ve.clicked.connect(lambda: self._in_ve(dialog, bien_lai))

        xe_cho = bien_lai.get("xe_vao_tu_hang_doi")
        if xe_cho is not None:
            QMessageBox.information(
                dialog,
                "Hàng đợi",
                f"Xe {xe_cho.bien_so} đã được tự động xếp vào vị trí {ma_vi_tri}.",
            )
        dialog.exec_()


    def _in_ve(self, parent, bien_lai):
        """In vé xe ra bằng QPrinter."""
        xe_ra = bien_lai["xe_ra"]
        ma_vi_tri = bien_lai["ma_vi_tri"]
        thoi_gian_ra = bien_lai["thoi_gian_ra"]
        tien_gui = bien_lai["tien_gui"]
        so_phut = int(bien_lai["so_phut_gui"])
        loai_xe = "Xe máy" if xe_ra.loai_xe == "XeMay" else "Ô tô"
        thanh_tien_str = f"{tien_gui:,}".replace(",", ".") + " đ"

        html = f"""
        <div style="font-family: Arial; width: 300px; margin: auto; padding: 20px;">
            <h2 style="text-align: center; border-bottom: 2px dashed #333; padding-bottom: 10px;">
                🅿️ VÉ GỬI XE
            </h2>
            <p style="text-align: center; font-size: 12px; color: #666;">
                Hệ thống quản lý bãi xe thông minh
            </p>
            <hr style="border: 1px dashed #ccc;">
            <table style="width: 100%; font-size: 14px; border-collapse: collapse;">
                <tr><td style="padding: 6px 0; font-weight: bold;">Biển số:</td>
                    <td style="padding: 6px 0; text-align: right;">{xe_ra.bien_so}</td></tr>
                <tr><td style="padding: 6px 0; font-weight: bold;">Loại xe:</td>
                    <td style="padding: 6px 0; text-align: right;">{loai_xe}</td></tr>
                <tr><td style="padding: 6px 0; font-weight: bold;">Vị trí đỗ:</td>
                    <td style="padding: 6px 0; text-align: right;">{ma_vi_tri}</td></tr>
                <tr><td style="padding: 6px 0; font-weight: bold;">Giờ vào:</td>
                    <td style="padding: 6px 0; text-align: right;">{xe_ra.thoi_gian_vao.strftime("%d/%m/%Y %H:%M")}</td></tr>
                <tr><td style="padding: 6px 0; font-weight: bold;">Giờ ra:</td>
                    <td style="padding: 6px 0; text-align: right;">{thoi_gian_ra.strftime("%d/%m/%Y %H:%M")}</td></tr>
                <tr><td style="padding: 6px 0; font-weight: bold;">Thời gian gửi:</td>
                    <td style="padding: 6px 0; text-align: right;">{so_phut // 60} giờ {so_phut % 60} phút</td></tr>
            </table>
            <hr style="border: 1px dashed #ccc;">
            <h3 style="text-align: center; color: #EA580C;">
                Thành tiền: {thanh_tien_str}
            </h3>
            <p style="text-align: center; font-size: 11px; color: #999; margin-top: 16px;">
                Cảm ơn quý khách đã sử dụng dịch vụ!
            </p>
        </div>
        """

        printer = QPrinter(QPrinter.HighResolution)
        print_dialog = QPrintDialog(printer, parent)
        if print_dialog.exec_() == QPrintDialog.Accepted:
            doc = QTextDocument()
            doc.setHtml(html)
            doc.print_(printer)
            QMessageBox.information(parent, "Thành công", "Đã in vé thành công!")


    # ----------------------------------------------------------------
    # Task 3: Xuất file Excel thống kê
    # ----------------------------------------------------------------
    def _xuat_excel_thong_ke(self, parent):
        """Xuất file Excel chứa doanh thu và xe gửi lâu nhất."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.warning(
                parent,
                "Thiếu thư viện",
                "Cần cài đặt openpyxl để xuất Excel.\n"
                "Chạy lệnh: pip install openpyxl",
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            "Lưu file thống kê",
            "thong_ke_bai_xe.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        wb = openpyxl.Workbook()

        # --- Style ---
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="176BFF", end_color="176BFF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        cot_headers = ["Biển số", "Loại xe", "Vị trí", "Giờ vào", "Giờ ra", "Số phút gửi", "Thành tiền"]
        cot_keys = ["bien_so", "loai_xe", "vi_tri", "gio_vao", "gio_ra", "so_phut_gui", "thanh_tien"]

        def ghi_sheet(ws, tieu_de, du_lieu):
            ws.title = tieu_de
            # Ghi tiêu đề sheet
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cot_headers))
            ws.cell(row=1, column=1, value=tieu_de).font = Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            # Ghi header
            for col_idx, header in enumerate(cot_headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Ghi dữ liệu
            for row_idx, ban_ghi in enumerate(du_lieu, 4):
                for col_idx, key in enumerate(cot_keys, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=ban_ghi.get(key, ""))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            # Auto width
            for col_idx in range(1, len(cot_headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

        # Sheet 1: Doanh thu cao nhất
        ws1 = wb.active
        du_lieu_doanh_thu = self._bai_xe.lay_lich_su_doanh_thu_cao_nhat()
        ghi_sheet(ws1, "Doanh thu", du_lieu_doanh_thu)

        # Sheet 2: Xe gửi lâu nhất
        ws2 = wb.create_sheet()
        du_lieu_gui_lau = self._bai_xe.lay_lich_su_gui_lau_nhat()
        ghi_sheet(ws2, "Xe gửi lâu nhất", du_lieu_gui_lau)

        try:
            wb.save(file_path)
            QMessageBox.information(parent, "Thành công", f"Đã xuất Excel:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(parent, "Lỗi", f"Không thể lưu file:\n{str(e)}")


    def _cap_nhat_cac_bang_quan_ly(self):
        for window in self._windows:
            if hasattr(window, "tableWidget") and hasattr(window, "tableWidget_2"):
                self._cap_nhat_bang_quan_ly(window)


    def _cap_nhat_bang_quan_ly(self, window):
        self._do_du_lieu_bang(
            window.tableWidget,
            self._bai_xe.lay_danh_sach_xe_dang_gui(),
            ["bien_so", "loai_xe", "vi_tri", "gio_vao", "hanh_dong"],
            gia_tri_mac_dinh={"hanh_dong": "CHECK-OUT"},
        )
        self._do_du_lieu_bang(
            window.tableWidget_2,
            self._bai_xe.lay_lich_su_xe_ra(),
            ["bien_so", "loai_xe", "vi_tri", "gio_vao", "gio_ra", "thanh_tien"],
        )


    def _do_du_lieu_bang(self, table, rows, columns, gia_tri_mac_dinh=None):
        gia_tri_mac_dinh = gia_tri_mac_dinh or {}
        table.setRowCount(len(rows))
        for row_index, row_data in enumerate(rows):
            for col_index, column in enumerate(columns):
                value = row_data.get(column, gia_tri_mac_dinh.get(column, ""))
                table.setItem(row_index, col_index, QTableWidgetItem(str(value)))
        table.resizeColumnsToContents()
